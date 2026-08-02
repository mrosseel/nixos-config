"""Nightly one-way export of the trip plan to Google Drive.

The git remote is the backup that matters — it has full history and it is what
you restore from. This exists for a different reason: the plan has to stay
readable by someone who does not use git, in the place they already look, if
the site is ever down. So it writes a formatted workbook into the same Drive
folder as the planning doc, and rclone hands it to Drive as a *native Google
Sheet* rather than a file you have to download to read.

One-way, always. Anything typed into the Sheet is overwritten the next night;
the sheet says so in its first cell.

Needs, both placed once by hand and not in the nix store:
  /var/lib/thailand-planner-secrets/rclone.conf   (rclone config, remote 'gdrive')
"""

import json
import os
import re
import subprocess
import sys
import tempfile
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PLAN = os.environ.get('PLAN_FILE', '/var/lib/thailand-planner/plan.json')
RCLONE_CONF = os.environ.get('RCLONE_CONF', '/var/lib/thailand-planner-secrets/rclone.conf')
REMOTE = os.environ.get('DRIVE_REMOTE', 'gdrive:Thailand')
SHEET_NAME = os.environ.get('DRIVE_SHEET_NAME', 'Reisplan (automatisch)')

TAG_RE = re.compile(r'<[^>]+>')
BREAK_RE = re.compile(r'</(p|div|li|ul|ol|h3|h4|blockquote)>|<br\s*/?>', re.I)


def plain(v):
    """Notes are rich text in the app; a spreadsheet cell wants words."""
    if not v:
        return ''
    s = BREAK_RE.sub('\n', str(v))
    s = TAG_RE.sub('', s)
    for a, b in (('&amp;', '&'), ('&lt;', '<'), ('&gt;', '>'), ('&nbsp;', ' '), ('&quot;', '"')):
        s = s.replace(a, b)
    return s.strip()


DOW_NL = ['maandag', 'dinsdag', 'woensdag', 'donderdag', 'vrijdag', 'zaterdag', 'zondag']

HEAD = PatternFill('solid', fgColor='1B2B47')
BAND = PatternFill('solid', fgColor='EEF2F9')
WARN = PatternFill('solid', fgColor='FFE9E6')
PAID = PatternFill('solid', fgColor='E4F5EE')
HEAD_FONT = Font(color='FFFFFF', bold=True)


def dates_for(plan):
    y, m, d = (int(x) for x in plan.get('start', '2026-10-21').split('-'))
    start = date(y, m, d)
    return [start + timedelta(days=i) for i in range(len(plan['days']))]


def style_header(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[ws.min_row]:
        c.fill = HEAD
        c.font = HEAD_FONT
        c.alignment = Alignment(vertical='center')
    ws.freeze_panes = ws.cell(row=ws.min_row + 1, column=1)


def sheet_plan(wb, plan, dates, stamp):
    ws = wb.active
    ws.title = 'Reisplan'
    ws.append([f'Automatische export van thailand.miker.be — laatst bijgewerkt {stamp}. '
               'Wijzigingen die je hier maakt gaan verloren; pas het plan aan op de site.'])
    ws.merge_cells('A1:G1')
    ws['A1'].font = Font(italic=True, color='8A4B3C')
    ws['A1'].fill = WARN
    ws.append(['Datum', 'Dag', 'Plaats', 'Wat', 'Weer', 'Overnachting', 'Status'])

    band = False
    prev = None
    for day, dt in zip(plan['days'], dates):
        place = day.get('place') or ''
        if place != prev:
            band = not band
            prev = place
        stays = [n for n in day.get('nodes') or [] if n.get('type') == 'hotel']
        stay = stays[0] if stays else None
        data = (stay or {}).get('data') or {}
        ws.append([
            dt.strftime('%d/%m/%Y'),
            DOW_NL[dt.weekday()],
            place or '(vrije dag)',
            ('★ ' if day.get('special') else '') + plain(day.get('notes')),
            day.get('weather') or '',
            data.get('name') or (stay or {}).get('title') or '',
            (stay or {}).get('status') or '',
        ])
        row = ws[ws.max_row]
        for c in row:
            c.alignment = Alignment(vertical='top', wrap_text=True)
            if band:
                c.fill = BAND
        if (stay or {}).get('status') == 'paid':
            row[6].fill = PAID

    for i, w in enumerate([12, 11, 22, 62, 26, 30, 10], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[2]:
        c.fill = HEAD
        c.font = HEAD_FONT
    ws.freeze_panes = 'A3'


def sheet_bookings(wb, plan, dates):
    ws = wb.create_sheet('Boekingen')
    ws.append(['Datum', 'Plaats', 'Soort', 'Wat', 'Van', 'Tot', 'Prijs', 'Geboekt door',
               'Betaald via', 'Annuleren vóór', 'Bevestiging', 'Status', 'Link'])
    for day, dt in zip(plan['days'], dates):
        for n in day.get('nodes') or []:
            if n.get('type') not in ('hotel', 'transport'):
                continue
            d = n.get('data') or {}
            ws.append([
                dt.strftime('%d/%m/%Y'),
                day.get('place') or '',
                'Overnachting' if n['type'] == 'hotel' else 'Vervoer',
                d.get('name') or n.get('title') or d.get('mode') or '',
                d.get('checkIn') or d.get('depart') or '',
                d.get('checkOut') or d.get('arrive') or '',
                d.get('price') or '',
                d.get('bookedBy') or '',
                d.get('paidVia') or '',
                d.get('cancelBefore') or '',
                d.get('confirmation') or d.get('ref') or '',
                n.get('status') or '',
                d.get('url') or d.get('bookingUrl') or '',
            ])
            if n.get('status') == 'paid':
                for c in ws[ws.max_row]:
                    c.fill = PAID
    style_header(ws, [12, 18, 14, 34, 12, 12, 12, 14, 14, 14, 18, 10, 44])


def sheet_notes(wb, plan, dates):
    ws = wb.create_sheet('Notities')
    ws.append(['Datum', 'Plaats', 'Soort', 'Notitie'])
    for day, dt in zip(plan['days'], dates):
        for n in day.get('nodes') or []:
            if n['type'] in ('hotel', 'transport'):
                continue
            d = n.get('data') or {}
            ws.append([dt.strftime('%d/%m/%Y'), day.get('place') or '',
                       n.get('title') or n['type'],
                       plain(d.get('notes')) or d.get('url') or ''])
    if plan.get('notes'):
        ws.append([])
        ws.append(['', '', 'Planning', ''])
        for note in plan['notes']:
            ws.append(['', '', '', plain(note)])
    style_header(ws, [12, 18, 22, 110])
    for row in ws.iter_rows(min_row=2):
        row[3].alignment = Alignment(vertical='top', wrap_text=True)


def main():
    with open(PLAN) as f:
        plan = json.load(f)
    if not plan.get('days'):
        print('no days in plan, nothing to export', file=sys.stderr)
        return 1

    dates = dates_for(plan)
    stamp = plan.get('savedAt', '')
    wb = Workbook()
    sheet_plan(wb, plan, dates, stamp)
    sheet_bookings(wb, plan, dates)
    sheet_notes(wb, plan, dates)

    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, f'{SHEET_NAME}.xlsx')
        wb.save(path)
        # --drive-import-formats xlsx makes Drive store this as a real Google
        # Sheet, not an attachment you have to download before you can read it.
        subprocess.run(
            ['rclone', '--config', RCLONE_CONF, '--drive-import-formats', 'xlsx',
             'copy', path, REMOTE],
            check=True, timeout=300)
    print(f'exported {len(plan["days"])} days to {REMOTE}/{SHEET_NAME}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
